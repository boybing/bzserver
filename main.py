import time,random,ssl
from urllib import request
from urllib.parse import quote
from bs4 import BeautifulSoup
from openpyxl import Workbook
import mongo as db


def reduceSame(vl):
    arr=[]
    for id in vl:
        if id not in arr:
            arr.append(id)
    # print(arr)
    return arr

def reqst(str, fg, count):
    if fg==1:
        req = request.Request(quote(str, safe='/:?='))
    else:
        req = request.Request(str)
    req.add_header('User-Agent',
                   'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6')
    with request.urlopen(req) as response:
        html = response.read().decode("utf-8")

    bs = BeautifulSoup(html, "html.parser")
    dd = bs.find_all(name='div',attrs={"class":"txt"})
    ee = bs.find_all(name='div', attrs={"class": "comment"})
    ff = bs.find_all(name='span',attrs={"class":"addr"})
    gg = bs.find_all(name='a',attrs={"class":"mean-price"})

    for i in range(len(gg)):
        try:
            if gg[i].b is None:
                vl.append({'title': dd[i].a.h4.string, 'price': "￥0", 'stars': ee[i].span.get('title'),
                           'address': ff[i].string, "review": ee[i].a.b.string,"url":dd[i].a.get('href'),"time":time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))})
            else:
                vl.append({'title': dd[i].a.h4.string, 'price': gg[i].b.string, 'stars': ee[i].span.get('title'),
                           'address': ff[i].string, "review": ee[i].a.b.string,"url":dd[i].a.get('href'),"time":time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))})
        except:
            vl.append({'title': dd[i].a.h4.string, 'price': "￥0", 'stars': ee[i].span.get('title'),
                       'address': ff[i].string, "review": "0","url":dd[i].a.get('href'),"time":time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))})
    c=bs.find(name='a', attrs={"class": "next"})

    try:
        time.sleep(random.randint(0, 1000) * 0.001)
        print("已完成第%i页数据添加"%count)
        count=count+1
        reqst(c.get('href'),0,count)
    except AttributeError as e:
        # print(reduceSame(vl))
        # db.insert(vl)
        db.insert(vl)
        db.find()
        # wb = Workbook()
        # ws = []
        # ws.append(wb.create_sheet(title='单页搜索结果'))
        # ws[0].append(['商家', '人均单价', '星级', '地址', '点评数',"数据更新时间"])
        # for i in reduceSame(vl):
        #     ws[0].append(['=HYPERLINK("' + i['url'] + '","' + i['title'] + '")',i['price'],i['stars'],i['address'],i['review'],i['time']])
        # save_path = "搜索结果.xlsx"
        # wb.save(save_path)


ssl._create_default_https_context = ssl._create_unverified_context
str="http://www.dianping.com/search/keyword/3/0_mamala"
vl = []
# print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))

reqst(str,1,1)

# 附带不转换字符参数
# print('\n附加不转换字符参数：\n%s' % quote(str, safe='/:?='))
    # str = re.sub(r'%25+', '%', str)

# from urllib import request
#
# with request.urlopen('https://api.douban.com/v2/book/2129650') as f:
#     data = f.read()
#     print('Status:', f.status, f.reason)
#     for k, v in f.getheaders():
#         print('%s: %s' % (k, v))
#     print('Data:', data.decode('utf-8'))

# wd=wordDeliver.deliver("线程是程序执行时的最小单位，它是进程的一个执行流，\
# #         是CPU调度和分派的基本单位，一个进程可以由很多个线程组成，\
# #         线程间共享进程的所有资源，每个线程有自己的堆栈和局部变量。\
# #         线程由CPU独立调度执行，在多CPU环境下就允许多个线程同时运行。\
# #         同样多线程也可以实现并发操作，每个请求分配一个线程来处理。")
# print(wd.data)


# t1=['a','b']
# t2=['c']
# t1.append(t2)
# print(t1)
# print('c' in t1)